"""
export_helper.py — Export mensual a Excel: P&L + Gastos + Facturacion.
"""
import io
from datetime import date


def generar_excel_mensual(mes: int, año: int) -> bytes:
    """Workbook con 3 hojas: P&L por area, Gastos del mes, Facturacion por cliente."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    from modulo_gastos import (_leer_gastos, _cargar_config, _finanzas_detallado,
                                _GASTOS_VEGGI_MAP, _VEGGI_RIO_PCT, _VEGGI_ANT_PCT,
                                _VEGGI_CHIM_PCT, _VEGGI_HOG_PCT)
    from excel_helper import leer_pedidos
    from data_helper  import cargar_clientes
    from config       import calcular_liquido

    cfg        = _cargar_config()
    campo_clis = cfg["campo_clis"]
    pedidos    = leer_pedidos()
    gastos_all = _leer_gastos()

    filtro_mes = lambda p: (p["fecha"] and p["fecha"].month == mes
                            and p["fecha"].year == año)
    gas_mes    = [g for g in gastos_all
                  if g["fecha"] and g["fecha"].month == mes and g["fecha"].year == año]

    # Zona map
    cli_zona = {}
    for c in cargar_clientes():
        for z, cods in _GASTOS_VEGGI_MAP.items():
            if c.get("codigo_lugar","") in cods:
                cli_zona[c["nombre"].lower().strip()] = z
                break

    fin = _finanzas_detallado(pedidos, campo_clis, filtro_mes, cli_zona)
    inc = fin["inc"]; costo_p = fin["costo"]
    _t  = lambda d: sum(d.values())

    def _gas_cat(cat):
        return sum(g["monto"] for g in gas_mes if g["categoria"] == cat)
    gas_campo_t = _gas_cat("Campo")
    gas_veggi_t = _gas_cat("Veggi")

    # ── Estilos ───────────────────────────────────────────────────────────────
    wb = Workbook()
    H  = Font(bold=True, color="FFFFFF")
    HF = PatternFill("solid", fgColor="2D7A2D")
    B  = Font(bold=True)
    TH = Border(top=Side(style="thin"))
    MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
             "Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

    def _header(ws, cols, widths):
        for j, (col, w) in enumerate(zip(cols, widths), start=1):
            cell = ws.cell(row=1, column=j, value=col)
            cell.font = H; cell.fill = HF
            ws.column_dimensions[cell.column_letter].width = w

    # ── Hoja 1: P&L ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "P&L"
    _header(ws, ["Area","Ingreso","Costo Producto","Gastos Op","Margen Neto"],
            [24, 14, 16, 13, 14])

    areas = [
        ("Campo",               _t(inc["Campo"]), 0,
         gas_campo_t),
        ("Veggi Rio",           _t(inc["Veggi"]["Rio"]),
         _t(costo_p["Veggi"]["Rio"]),  round(gas_veggi_t*_VEGGI_RIO_PCT, 2)),
        ("Veggi Antigua",       _t(inc["Veggi"]["Antigua"]),
         _t(costo_p["Veggi"]["Antigua"]), round(gas_veggi_t*_VEGGI_ANT_PCT, 2)),
        ("Veggi Chimaltenango", _t(inc["Veggi"]["Chimaltenango"]),
         _t(costo_p["Veggi"]["Chimaltenango"]), round(gas_veggi_t*_VEGGI_CHIM_PCT, 2)),
        ("Veggi Hogares",       _t(inc["Interno"]),
         _t(costo_p["Interno"]), round(gas_veggi_t*_VEGGI_HOG_PCT, 2)),
    ]
    r = 2
    tot = [0.0, 0.0, 0.0, 0.0]
    for nombre, ing, cc, gas in areas:
        mn = ing - cc - gas
        for j, v in enumerate([nombre, ing, cc, gas, mn], start=1):
            ws.cell(row=r, column=j, value=v)
        for j, v in enumerate([ing, cc, gas, mn]):
            tot[j] += v
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = B
    for j, v in enumerate(tot, start=2):
        cl = ws.cell(row=r, column=j, value=round(v, 2)); cl.font = B; cl.border = TH
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cl in row: cl.number_format = "#,##0.00"

    # ── Hoja 2: Gastos ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Gastos")
    _header(ws2, ["Fecha","Semana","Categoria","SubCategoria","Area",
                  "Frecuencia","Proveedor","Concepto","Monto"],
            [11, 8, 11, 18, 13, 11, 16, 26, 11])
    r = 2
    for g in sorted(gas_mes, key=lambda x: x["fecha"] or date.min):
        vals = [g["fecha"].strftime("%d/%m/%Y") if g["fecha"] else "",
                g["semana"], g["categoria"], g["subcat"], g["area"],
                g["frecuencia"], g["proveedor"], g["concepto"], g["monto"]]
        for j, v in enumerate(vals, start=1):
            ws2.cell(row=r, column=j, value=v)
        r += 1
    ws2.cell(row=r, column=8, value="TOTAL").font = B
    cl = ws2.cell(row=r, column=9, value=round(sum(g["monto"] for g in gas_mes), 2))
    cl.font = B; cl.border = TH
    for row in ws2.iter_rows(min_row=2, min_col=9, max_col=9):
        for c2 in row: c2.number_format = "#,##0.00"

    # ── Hoja 3: Facturacion por cliente ───────────────────────────────────────
    ws3 = wb.create_sheet("Facturacion")
    _header(ws3, ["Cliente","Total","Base IVA","ISR","Descuento","Liquido"],
            [26, 13, 13, 11, 11, 13])
    por_cli: dict = {}
    for p in pedidos:
        if not filtro_mes(p): continue
        if p["status"] == "Cancelado": continue
        por_cli[p["cliente"]] = por_cli.get(p["cliente"], 0) + float(p.get("total") or 0)

    r = 2
    tots = [0.0]*5
    for cli in sorted(por_cli, key=lambda x: -por_cli[x]):
        total = por_cli[cli]
        liq, isr, desc = calcular_liquido(cli, total)
        base = round(total / 1.12, 2)
        vals = [cli, total, base, isr, desc, liq]
        for j, v in enumerate(vals, start=1):
            ws3.cell(row=r, column=j, value=v)
        for j, v in enumerate(vals[1:]):
            tots[j] += v
        r += 1
    ws3.cell(row=r, column=1, value="TOTAL").font = B
    for j, v in enumerate(tots, start=2):
        cl = ws3.cell(row=r, column=j, value=round(v, 2)); cl.font = B; cl.border = TH
    for row in ws3.iter_rows(min_row=2, min_col=2):
        for c3 in row: c3.number_format = "#,##0.00"

    # Titulo en propiedades
    wb.properties.title = f"VeggiExpress {MESES[mes-1]} {año}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
